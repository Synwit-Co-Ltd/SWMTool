#! python3
import re
import sys
import openpyxl as xl


def parse(xlsx, sheet, pad_column, pin_column, chip):
    wb = xl.open(xlsx)
    ws = wb[sheet]
    
    pins = {}
    for row in range(1, 256):   # 1-based index
        pad = ws[f'{pad_column}{row}'].value

        if not isinstance(ws[f'{pin_column}{row}'], xl.cell.MergedCell):
            pin = ws[f'{pin_column}{row}'].value

        if type(pin) == int:
            if pin not in pins:
                pins[pin] = []

            pins[pin].append(re.sub(r'(^[A-FMNP]\d+$)', r'P\1', pad.replace('pad_', '').upper()))

    for pin, pad in pins.items():
        pad.sort(key=lambda x: x[1:] if x[0] == 'P' else x)

    with open(f'{chip}.txt', 'w', encoding='utf-8') as f:
        for pin in sorted(pins):
            f.write(f'{pin:2d}: {" ".join(pins[pin])}\n')


parse(r'D:\synwit_doc\odyssey_2009\11_芯片封装\MP\LQFP64\SWM3410S_MP_Bonding_information_LQFP64.xlsx', 'SWM3410MP', 'D', 'I', 'SWM34SRxTy')
parse(r'D:\synwit_doc\odyssey_2009\11_芯片封装\MP\LQFP100\SWM3410_MP_Bonding_information_LQFP100.xlsx', 'SWM3410MPW2', 'D', 'I', 'SWM341VxTy')
parse(r'D:\synwit_doc\odyssey_2009\11_芯片封装\MP\LQFP100\SWM3410S_MP_Bonding_information_LQFP100V3.xlsx', 'SWM3410MP', 'D', 'I', 'SWM34SVxTy')
